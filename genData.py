# -*- coding: utf-8 -*-

import os
from docx import Document
import openpyxl

file_path = r'E:\temp\走出去调查表\参股外国企业信息表\信息表'
infoWorkbook_path = r'E:\temp\走出去调查表\参股外国企业信息表\参股外国企业信息表.xlsm'
infoWorkbook = openpyxl.load_workbook(infoWorkbook_path)
outPath = r'E:\temp\走出去调查表\参股外国企业信息表\result'
if not os.path.exists(outPath):
    os.mkdir(outPath)

for file in os.listdir(file_path):
    print('正在处理文件：%s'%file)
    temp = file[14:]
    corpName = temp.split('.')[0]
    infoSheet = infoWorkbook[corpName]
    doc = Document(os.path.join(file_path,file))
    table = doc.tables[0]
    table.cell(1,2).text = infoSheet['D4'].value  # 报告公司名称
    table.cell(1,9).text = infoSheet['K4'].value if infoSheet['K4'].value else '' # 报告公司纳税人识别号
    table.cell(3,2).text = infoSheet['D6'].value # 被投资外国企业名称
    if infoSheet['K6'].value:
        table.cell(3,9).text = infoSheet['K6'].value # 被投资外国企业纳税人识别号
    else:
        table.cell(3,9).text = ''
    table.cell(4,2).text = infoSheet['D7'].value if infoSheet['D7'].value else ''# 被投资方成立地
    table.cell(4,9).text = infoSheet['K7'].value if infoSheet['K7'].value else ''# 被投资方主营业务类型
    table.cell(5,2).text = '%.2f%%'%(infoSheet['D8'].value*100) # 报告人持股比例
    for i in [8,9,14]:
        for j in [0,1,4,8]:
            temp = infoSheet.cell(row=i+3,column=j+2).value
            v = temp if temp else ''
            if isinstance(v,float):
                v = '%.2f%%'%(v*100)
            table.cell(i,j).text = v
    doc.save(os.path.join(outPath,file))
    print("文件'%s'处理完毕"%file)

infoWorkbook.close()
